from main_file import run_sql
import openpyxl
import os
from sqlalchemy import text
from datetime import datetime


def create_processed_nreading(conn):
    os.makedirs("data/processed", exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_name = conn.execute(text("""
            SELECT station_id,name
            FROM stations
    """)).fetchall()

    for station_id, station_name in sheet_name:
        sheet = wb.create_sheet(title=station_name)
        sheet.append(["ts_utc", "db_level", "part_of_day", "src_month"])

        rows = conn.execute(
            text("""
                SELECT ts_utc, db_level, part_of_day, src_month
                FROM noise_reading
                WHERE station_id = :station_id
                ORDER BY reading_id;
            """),
            {"station_id": station_id}
        ).fetchall()

        for ro in rows:
            cleaned = []
            for val in ro:
                if isinstance(val, datetime) and val.tzinfo is not None:
                    val = val.replace(tzinfo=None)
                cleaned.append(val)
            sheet.append(cleaned)

    try:
        wb.save("data/processed/nr_processed.xlsx")
        print("File successfuly made : data/processed/nr_processed.xlsx")
    except PermissionError:
        print("Already open in Excel.")


def create_processed_level_l(conn):
    os.makedirs("data/processed", exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_name = conn.execute(text("""
            SELECT station_id,name
            FROM stations
    """)).fetchall()

    for station_id, station_name in sheet_name:
        sheet = wb.create_sheet(title=station_name)

        sheet.append(["ts_hour_kst", "n_samples", "laeq"])

        rows = conn.execute(
            text("""
                    SELECT ts_hour_kst, n_samples, laeq
                    FROM noise_level_h
                    WHERE station_id = :station_id
                    ORDER BY station_id;
                """),
            {"station_id": station_id}
        ).fetchall()

        for ro in rows:
            sheet.append(tuple(ro))

    try:
        wb.save("data/processed/noise_level_l.xlsx")
        print("File successfuly made : data/processed/noise_level_l.xlsx")
    except PermissionError:
        print("Already open in Excel.")


def create_processed_level_d(conn):
    os.makedirs("data/processed", exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_name = conn.execute(text("""
            SELECT station_id,name
            FROM stations
    """)).fetchall()

    for station_id, station_name in sheet_name:
        sheet = wb.create_sheet(title=station_name)

        sheet.append(["d_kst", "laeq_day", "laeq_night"])

        rows = conn.execute(
            text("""
                    SELECT d_kst, laeq_day, laeq_night
                    FROM noise_level_d
                    WHERE station_id = :station_id
                    ORDER BY station_id;
                """),
            {"station_id": station_id}
        ).fetchall()

        for ro in rows:
            sheet.append(tuple(ro))

    try:
        wb.save("data/processed/noise_level_d.xlsx")
        print("File successfuly made : data/processed/noise_level_d.xlsx")
    except PermissionError:
        print("Already open in Excel.")


def create_processed_peak_time(conn):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    stations = conn.execute(text("""
        SELECT station_id, name
        FROM stations
    """)).fetchall()

    from main_file import fetch_peak_times
    day_peak_df, global_peak_df = fetch_peak_times(conn)

    for station_id, station_name in stations:
        sheet = wb.create_sheet(title=station_name)
        sheet.append(["date", "hour", "laeq", "peak_type"])

        day_rows = day_peak_df[day_peak_df["station_id"] == station_id]
        for _, row in day_rows.iterrows():
            sheet.append([
                row["d_kst"],
                row["hour_kst"],
                row["laeq"],
                "day_peak"
            ])

        global_rows = global_peak_df[global_peak_df["station_id"]
                                     == station_id]
        for _, row in global_rows.iterrows():
            sheet.append([
                row["d_kst"],
                row["hour_kst"],
                row["laeq"],
                "global_peak"
            ])

    try:
        wb.save("data/processed/peak_time.xlsx")
        print("File successfuly made : data/processed/peak_time.xlsx")
    except PermissionError:
        print("Already open in Excel.")


if __name__ == "__main__":
    from main_file import connect_engine
    engine = connect_engine()
    with engine.begin() as conn:
        print("→ Creating processed tables...")
        create_processed_nreading(conn)
        create_processed_level_l(conn)
        create_processed_level_d(conn)
        create_processed_peak_time(conn)
        print("Done.")
